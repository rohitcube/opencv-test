
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import os
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors


save_path = '/Users/rohit/Downloads/'
file_name = 'query_comparison_2.xlsx'

wb = load_workbook(os.path.join(save_path, file_name))
ws = wb.active

#Custom
bold_font = Font(bold=True)
for column in ['A', 'B', 'C']:
    for cell in ws[column]:
        cell.alignment = Alignment(wrapText=True)

# Instructions

ws['A1'] = 'QUERY'
ws['A1'].font = bold_font
ws['B1'] = 'TABLE NAME'
ws['B1'].font = bold_font
ws['B1'] = 'TABLE TYPE'
ws['B1'].font = bold_font
ws['D1'] = 'QUERY TIMING'
ws['D1'].font = bold_font
ws['E1'] = 'QUERY ID'
ws['E1'].font = bold_font

text = """

SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-02-23' AND '2024-03-03' AND TOKEN_ID IN ('77000203861382602893', '77000275044902663753', '77000255653273996437', '77000221783810232610');
01b4d1e3-3201-1a18-0003-0d6202f838be
Success
180ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-02-23' AND '2024-03-03' AND TOKEN_ID IN ('77000203861382602893', '77000275044902663753', '77000255653273996437', '77000221783810232610');
01b4d1e3-3201-1a18-0003-0d6202f838b6
Success
4.0s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-02-23' AND '2024-03-03' AND TOKEN_ID IN ('77000203861382602893', '77000275044902663753', '77000255653273996437', '77000221783810232610');
01b4d1e3-3201-1a18-0003-0d6202f838ae
Success
2.6s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e3-3201-1a18-0003-0d6202f838aa
Success
28ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000287968460428239', '77000293345129123135', '77000255653273996437', '77000221783810232610');
01b4d1e3-3201-1ac3-0003-0d6202f85796
Success
416ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000287968460428239', '77000293345129123135', '77000255653273996437', '77000221783810232610');
01b4d1e2-3201-1abd-0003-0d6202f8482a
Success
12s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000287968460428239', '77000293345129123135', '77000255653273996437', '77000221783810232610');
01b4d1e2-3201-1ac3-0003-0d6202f8578e
Success
3.0s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1ac3-0003-0d6202f85786
Success
24ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000217800973247507', '77000283904747063908', '77000265246810428769', '77000201399325240887');
01b4d1e2-3201-1abd-0003-0d6202f84822
Success
1.9s
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000217800973247507', '77000283904747063908', '77000265246810428769', '77000201399325240887');
01b4d1e2-3201-1a18-0003-0d6202f83892
Success
9.5s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-03-07' AND '2024-03-22' AND TOKEN_ID IN ('77000217800973247507', '77000283904747063908', '77000265246810428769', '77000201399325240887');
01b4d1e2-3201-1ac3-0003-0d6202f85766
Success
1.2s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1ac3-0003-0d6202f8575e
Success
22ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-07' AND '2024-04-19' AND TOKEN_ID IN ('77000242651920605554', '77000284658053120226', '77000250932605697361');
01b4d1e2-3201-1a18-0003-0d6202f83876
Success
3.8s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-07' AND '2024-04-19' AND TOKEN_ID IN ('77000242651920605554', '77000284658053120226', '77000250932605697361');
01b4d1e2-3201-1a18-0003-0d6202f83872
Success
842ms
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1abd-0003-0d6202f847ba
Success
34ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-07' AND '2024-04-19' AND TOKEN_ID IN ('77000218855433140455', '77000292082324526254', '77000270138743715842');
01b4d1e2-3201-1a18-0003-0d6202f8386e
Success
4.1s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-07' AND '2024-04-19' AND TOKEN_ID IN ('77000218855433140455', '77000292082324526254', '77000270138743715842');
01b4d1e2-3201-1a18-0003-0d6202f8386a
Success
634ms
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1a18-0003-0d6202f83866
Success
22ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-04' AND '2024-04-11' AND TOKEN_ID IN ('77000263702561289791', '77000279711242413995', '77000212556083635732');
01b4d1e2-3201-1abd-0003-0d6202f847ae
Success
4.2s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-04' AND '2024-04-11' AND TOKEN_ID IN ('77000263702561289791', '77000279711242413995', '77000212556083635732');
01b4d1e2-3201-1ac3-0003-0d6202f85722
Success
470ms
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1a18-0003-0d6202f8385e
Success
25ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-04-04' AND '2024-04-11' AND TOKEN_ID IN ('77000221091517024469', '77000220838434381124', '77000212222026898384');
01b4d1e2-3201-1a18-0003-0d6202f8385a
Success
478ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-04' AND '2024-04-11' AND TOKEN_ID IN ('77000221091517024469', '77000220838434381124', '77000212222026898384');
01b4d1e2-3201-1a18-0003-0d6202f83856
Success
4.6s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-04' AND '2024-04-11' AND TOKEN_ID IN ('77000221091517024469', '77000220838434381124', '77000212222026898384');
01b4d1e2-3201-1a18-0003-0d6202f83852
Success
699ms
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e2-3201-1a18-0003-0d6202f8384e
Success
25ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000288963995490002', '77000289766404357683', '77000241767342090795');
01b4d1e2-3201-1ac3-0003-0d6202f8570e
Success
743ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000288963995490002', '77000289766404357683', '77000241767342090795');
01b4d1e1-3201-1a18-0003-0d6202f83846
Success
4.8s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000288963995490002', '77000289766404357683', '77000241767342090795');
01b4d1e1-3201-1ac3-0003-0d6202f8570a
Success
1.1s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e1-3201-1a18-0003-0d6202f83842
Success
31ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000273086285951897', '77000230194054772099', '77000259144328830912');
01b4d1e1-3201-1abd-0003-0d6202f84796
Success
1.0s
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000273086285951897', '77000230194054772099', '77000259144328830912');
01b4d1e1-3201-1a18-0003-0d6202f8383a
Success
5.3s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-04-01' AND '2024-04-15' AND TOKEN_ID IN ('77000273086285951897', '77000230194054772099', '77000259144328830912');
01b4d1e1-3201-1ac3-0003-0d6202f856da
Success
1.0s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e1-3201-1abd-0003-0d6202f84776
Success
38ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-02-14' AND '2024-03-01' AND TOKEN_ID IN ('77000241280660623543', '77000235491645978155', '77000244148540118580');
01b4d1e1-3201-1abd-0003-0d6202f84766
Success
8.2s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-02-14' AND '2024-03-01' AND TOKEN_ID IN ('77000241280660623543', '77000235491645978155', '77000244148540118580');
01b4d1e1-3201-1abd-0003-0d6202f84762
Success
792ms
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e1-3201-1a18-0003-0d6202f83826
Success
32ms
SELECT * FROM TX_ABT_MATCH_HYBRID_IDX WHERE BUSINESS_DATE BETWEEN '2024-02-14' AND '2024-03-01' AND TOKEN_ID IN ('23329022428530548075', '22482105290180707783', '77000227925421585405');
01b4d1e1-3201-1ac3-0003-0d6202f856ca
Success
448ms
SELECT * FROM TX_ABT_MATCH_HYBRID WHERE BUSINESS_DATE BETWEEN '2024-02-14' AND '2024-03-01' AND TOKEN_ID IN ('23329022428530548075', '22482105290180707783', '77000227925421585405');
01b4d1e0-3201-1ac3-0003-0d6202f85682
Success
1m 23s
SELECT * FROM TX_ABT_MATCH WHERE BUSINESS_DATE BETWEEN '2024-02-14' AND '2024-03-01' AND TOKEN_ID IN ('23329022428530548075', '22482105290180707783', '77000227925421585405');
01b4d1e0-3201-1abd-0003-0d6202f84732
Success
1.1s
ALTER SESSION SET USE_CACHED_RESULT = FALSE;
01b4d1e0-3201-1ac3-0003-0d6202f85672
Success
39ms

"""
lines = text.split('\n')

# Initialize an empty list to store the formatted data

row = 2

# Iterate through the lines
for i in range(len(lines)):
    line = lines[i].strip()
    if line.startswith('SELECT') or line.startswith('ALTER'):
        # Extract the query
        query = line

        # Extract table name
        query_list = query.split(' ')
        table_name = query_list[3].strip()

        # Table type
        table_type = " "
        if (table_name == 'TX_ABT_MATCH'):
            table_type = "Normal Table"
        elif (table_name == 'TX_ABT_MATCH_HYBRID'):
            table_type = "Hybrid Table"
        elif (table_name == 'TX_ABT_MATCH_HYBRID_IDX'):
            table_type = "Hybrid Table w/ Index"
        else:
           table_type = "err"

        # Extract the query ID
        query_id = lines[i + 1].strip()

        # Extract the timing
        timing = lines[i + 3].strip()

        # Write data to the ws
        ws.cell(row=row, column=1, value=query)
        ws.cell(row=row, column=2, value=table_name)
        ws.cell(row=row, column=3, value=table_type)
        ws.cell(row=row, column=4, value=timing)
        ws.cell(row=row, column=5, value=query_id)

        # Increment the row counter
        row += 1

"""
max_row = ws.max_row
# Iterate through the rows in reverse order
for row_num in range(max_row, 1, -1):
    # Check if the current row is a multiple of 3
    if (row_num - 1) % 3 == 0:
        # Insert a blank row after the current row
        ws.insert_rows(row_num)
"""

string_to_delete = "ALTER SESSION SET USE_CACHED_RESULT = FALSE;"
yellowFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FFFF00'))

"""
# Iterate through the rows
for cell in ws.iter_cols(min_col=4, max_col=4):
    # Get the cell value
    cell_value = cell[0].value
    if int(cell_value) > 5:
            # Set the fill color for the cell to yellow
            cell[0].fill = yellowFill
"""
try:
    wb.save(os.path.join(save_path, file_name))
except Exception as e:
    print(f"Error saving file: {e}")
else:
    print("File saved successfully")

